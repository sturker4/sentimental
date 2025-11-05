#!/usr/bin/env python3
"""
Convert a YC scraper checkpoint JSON into an Excel workbook.

The checkpoint produced by `scraper.py` stores a mapping of YC link -> field dict.
This utility flattens that structure into a tabular worksheet with one row per link
and the standard 10 columns expected by the scraper output.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover - dependency check
    raise SystemExit(
        "Missing dependency: openpyxl\n\n"
        "Install it with `python -m pip install openpyxl` (ideally inside your virtualenv) "
        "and rerun this command."
    ) from exc


COLUMNS: List[str] = [
    "YC Link",
    "Active Founders",
    "Founders LinkedIn Link",
    "Status",
    "Website",
    "Primary Partner",
    "Founded Year",
    "Team Size",
    "Batch",
    "Location",
]


def load_checkpoint(path: Path) -> Dict[str, Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8") as f:
            raw = json.load(f)
            if isinstance(raw, dict):
                return {str(k): dict(v) for k, v in raw.items() if isinstance(v, dict)}
    except Exception as exc:
        raise SystemExit(f"Failed to read checkpoint {path}: {exc}") from exc
    raise SystemExit(
        f"Unexpected checkpoint structure in {path}. Expected a JSON object."
    )


def iter_rows(data: Dict[str, Dict[str, Any]]) -> Iterable[List[Any]]:
    for url, fields in data.items():
        row = []
        for column in COLUMNS:
            if column == "YC Link":
                row.append(url)
            else:
                row.append(fields.get(column))
        yield row


def export_to_excel(data: Dict[str, Dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "YC Companies"
    ws.append(COLUMNS)
    for row in iter_rows(data):
        ws.append(row)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert scraper checkpoint JSON into Excel."
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to checkpoint JSON (e.g. yc_output.csv.ckpt.json)",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Destination .xlsx file (e.g. yc_output.xlsx)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    data = load_checkpoint(args.input)
    export_to_excel(data, args.output)
    print(f"Wrote {len(data)} rows to {args.output}")


if __name__ == "__main__":
    main()
